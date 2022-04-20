import tkinter as tk
from tkinter import Tk
from tkinter.messagebox import showwarning
import openpyxl
from openpyxl import Workbook
import win32com.client as win32
from time import sleep
#from tkinter import messagebox




# 实例化
wb = Workbook()
#激活 worksheet
ws = wb.active
# 保存表格
wb.save("test2222222.xlsx")

warn = lambda app: showwarning(app,'Exit?')
RANGE = range(3,8)

def excel():
    app = 'Excel'
    #xl = win32.gencache.EnsureDispatch('%s.Application' % app)
    xl = win32.Dispatch('%s.Application' % app)
    ss = xl.Workbooks.Add()
    sh = ss.ActiveSheet
    xl.Visible = True
    sleep(1)
    sh.Cells(1,1).Value = 'dasasdasd'
    warn(app)
    ss.Close(True)
    xl.Application.Quit()


def oprnexcel():
    filename = r'D:\wzk_python_selenium\PC4.0\files\yesOrNo.xlsx'
    wb = Workbook()
    # 激活 worksheet
    # ws = wb.active
    inwb = openpyxl.load_workbook(filename)
    sheets = inwb.sheetnames
    ws = inwb[sheets[0]]
    rows = ws.max_row
    cols = ws.max_column
    print(rows)
    print(cols)
    print(ws.cell(1, 1).value)

if __name__ == '__main__':
    # Tk().withdraw()
    # excel()
    oprnexcel()

